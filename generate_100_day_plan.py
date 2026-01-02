"""
Generate 100-Day DSA Preparation Plan as Excel Workbook

This script creates a comprehensive Excel workbook with:
- Daily schedule (100 days)
- LeetCode problems (250-300 problems)
- Revision calendar
- Progress tracking
- Topic distribution
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# LeetCode Problems Database
LEETCODE_PROBLEMS = {
    'arrays': [
        (217, 'Contains Duplicate', 'Easy', 'Hash Set'),
        (53, 'Maximum Subarray', 'Easy', 'Kadane\'s Algorithm'),
        (238, 'Product of Array Except Self', 'Medium', 'Prefix/Postfix'),
        (1, 'Two Sum', 'Easy', 'Hash Map'),
        (15, '3Sum', 'Medium', 'Two Pointers'),
        (11, 'Container With Most Water', 'Medium', 'Two Pointers'),
        (189, 'Rotate Array', 'Medium', 'Array Rotation'),
        (121, 'Best Time to Buy and Sell Stock', 'Easy', 'Greedy'),
        (42, 'Trapping Rain Water', 'Hard', 'Two Pointers/Stack'),
        (56, 'Merge Intervals', 'Medium', 'Sorting'),
        (48, 'Rotate Image', 'Medium', 'Matrix'),
        (54, 'Spiral Matrix', 'Medium', 'Matrix Traversal'),
        (73, 'Set Matrix Zeroes', 'Medium', 'Matrix'),
        (128, 'Longest Consecutive Sequence', 'Medium', 'Hash Set'),
        (287, 'Find the Duplicate Number', 'Medium', 'Floyd\'s Cycle'),
        (41, 'First Missing Positive', 'Hard', 'Array Manipulation'),
        (283, 'Move Zeroes', 'Easy', 'Two Pointers'),
        (88, 'Merge Sorted Array', 'Easy', 'Two Pointers'),
        (26, 'Remove Duplicates from Sorted Array', 'Easy', 'Two Pointers'),
        (80, 'Remove Duplicates from Sorted Array II', 'Medium', 'Two Pointers'),
    ],
    'linked_lists': [
        (206, 'Reverse Linked List', 'Easy', 'Reversal'),
        (141, 'Linked List Cycle', 'Easy', 'Floyd\'s Cycle'),
        (142, 'Linked List Cycle II', 'Medium', 'Floyd\'s Cycle'),
        (19, 'Remove Nth Node From End', 'Medium', 'Two Pointers'),
        (2, 'Add Two Numbers', 'Medium', 'Linked List Math'),
        (143, 'Reorder List', 'Medium', 'Reversal + Merge'),
        (21, 'Merge Two Sorted Lists', 'Easy', 'Merge'),
        (23, 'Merge k Sorted Lists', 'Hard', 'Heap/Merge'),
        (25, 'Reverse Nodes in k-Group', 'Hard', 'Reversal'),
        (138, 'Copy List with Random Pointer', 'Medium', 'Hash Map'),
        (92, 'Reverse Linked List II', 'Medium', 'Reversal'),
        (86, 'Partition List', 'Medium', 'Partition'),
        (148, 'Sort List', 'Medium', 'Merge Sort'),
        (234, 'Palindrome Linked List', 'Easy', 'Reversal'),
        (160, 'Intersection of Two Linked Lists', 'Easy', 'Two Pointers'),
    ],
    'strings': [
        (344, 'Reverse String', 'Easy', 'Two Pointers'),
        (125, 'Valid Palindrome', 'Easy', 'Two Pointers'),
        (3, 'Longest Substring Without Repeating', 'Medium', 'Sliding Window'),
        (5, 'Longest Palindromic Substring', 'Medium', 'DP/Expand'),
        (49, 'Group Anagrams', 'Medium', 'Hash Map'),
        (438, 'Find All Anagrams in a String', 'Medium', 'Sliding Window'),
        (76, 'Minimum Window Substring', 'Hard', 'Sliding Window'),
        (20, 'Valid Parentheses', 'Easy', 'Stack'),
        (151, 'Reverse Words in a String', 'Medium', 'String Manipulation'),
        (14, 'Longest Common Prefix', 'Easy', 'String Comparison'),
        (28, 'Find the Index of First Occurrence', 'Easy', 'String Search'),
        (459, 'Repeated Substring Pattern', 'Easy', 'String Pattern'),
        (424, 'Longest Repeating Character Replacement', 'Medium', 'Sliding Window'),
        (394, 'Decode String', 'Medium', 'Stack'),
        (227, 'Basic Calculator II', 'Medium', 'Stack'),
    ],
    'binary_search': [
        (704, 'Binary Search', 'Easy', 'Binary Search'),
        (35, 'Search Insert Position', 'Easy', 'Binary Search'),
        (162, 'Find Peak Element', 'Medium', 'Binary Search'),
        (33, 'Search in Rotated Sorted Array', 'Medium', 'Rotated Array'),
        (153, 'Find Minimum in Rotated Sorted Array', 'Medium', 'Rotated Array'),
        (34, 'Find First and Last Position', 'Medium', 'Binary Search'),
        (69, 'Sqrt(x)', 'Easy', 'Binary Search'),
        (74, 'Search a 2D Matrix', 'Medium', 'Binary Search'),
        (240, 'Search a 2D Matrix II', 'Medium', 'Binary Search'),
        (875, 'Koko Eating Bananas', 'Medium', 'Binary Search on Answer'),
        (1011, 'Capacity To Ship Packages', 'Medium', 'Binary Search on Answer'),
        (410, 'Split Array Largest Sum', 'Hard', 'Binary Search on Answer'),
        (4, 'Median of Two Sorted Arrays', 'Hard', 'Binary Search'),
        (81, 'Search in Rotated Sorted Array II', 'Medium', 'Rotated Array'),
        (154, 'Find Minimum in Rotated Sorted Array II', 'Hard', 'Rotated Array'),
    ],
    'sorting': [
        (912, 'Sort an Array', 'Medium', 'Sorting'),
        (75, 'Sort Colors', 'Medium', 'Three-way Partition'),
        (179, 'Largest Number', 'Medium', 'Custom Sort'),
        (315, 'Count of Smaller Numbers After Self', 'Hard', 'Merge Sort'),
        (493, 'Reverse Pairs', 'Hard', 'Merge Sort'),
        (215, 'Kth Largest Element in an Array', 'Medium', 'Quick Select'),
        (56, 'Merge Intervals', 'Medium', 'Sorting'),
        (57, 'Insert Interval', 'Medium', 'Sorting'),
        (252, 'Meeting Rooms', 'Easy', 'Sorting'),
        (253, 'Meeting Rooms II', 'Medium', 'Heap'),
        (164, 'Maximum Gap', 'Hard', 'Radix Sort'),
        (274, 'H-Index', 'Medium', 'Sorting'),
        (324, 'Wiggle Sort II', 'Medium', 'Sorting'),
    ],
    'stacks': [
        (20, 'Valid Parentheses', 'Easy', 'Stack'),
        (150, 'Evaluate Reverse Polish Notation', 'Medium', 'Stack'),
        (155, 'Min Stack', 'Medium', 'Stack Design'),
        (739, 'Daily Temperatures', 'Medium', 'Monotonic Stack'),
        (503, 'Next Greater Element II', 'Medium', 'Monotonic Stack'),
        (84, 'Largest Rectangle in Histogram', 'Hard', 'Monotonic Stack'),
        (85, 'Maximal Rectangle', 'Hard', 'Stack + DP'),
        (32, 'Longest Valid Parentheses', 'Hard', 'Stack'),
        (394, 'Decode String', 'Medium', 'Stack'),
        (227, 'Basic Calculator II', 'Medium', 'Stack'),
        (224, 'Basic Calculator', 'Hard', 'Stack'),
        (71, 'Simplify Path', 'Medium', 'Stack'),
    ],
    'queues': [
        (933, 'Number of Recent Calls', 'Easy', 'Queue'),
        (239, 'Sliding Window Maximum', 'Hard', 'Monotonic Queue'),
        (621, 'Task Scheduler', 'Medium', 'Greedy + Queue'),
        (346, 'Moving Average from Data Stream', 'Easy', 'Queue'),
        (362, 'Design Hit Counter', 'Medium', 'Queue'),
        (353, 'Design Snake Game', 'Medium', 'Queue'),
    ],
    'hash_tables': [
        (1, 'Two Sum', 'Easy', 'Hash Map'),
        (49, 'Group Anagrams', 'Medium', 'Hash Map'),
        (387, 'First Unique Character', 'Easy', 'Hash Map'),
        (560, 'Subarray Sum Equals K', 'Medium', 'Prefix Sum + Hash'),
        (3, 'Longest Substring Without Repeating', 'Medium', 'Hash Map'),
        (128, 'Longest Consecutive Sequence', 'Medium', 'Hash Set'),
        (36, 'Valid Sudoku', 'Medium', 'Hash Set'),
        (202, 'Happy Number', 'Easy', 'Hash Set'),
        (290, 'Word Pattern', 'Easy', 'Hash Map'),
        (205, 'Isomorphic Strings', 'Easy', 'Hash Map'),
        (454, '4Sum II', 'Medium', 'Hash Map'),
        (18, '4Sum', 'Medium', 'Two Pointers + Hash'),
    ],
    'bit_manipulation': [
        (191, 'Number of 1 Bits', 'Easy', 'Bit Counting'),
        (231, 'Power of Two', 'Easy', 'Bit Check'),
        (136, 'Single Number', 'Easy', 'XOR'),
        (137, 'Single Number II', 'Medium', 'Bit Manipulation'),
        (260, 'Single Number III', 'Medium', 'XOR'),
        (201, 'Bitwise AND of Numbers Range', 'Medium', 'Bit Manipulation'),
        (190, 'Reverse Bits', 'Easy', 'Bit Manipulation'),
        (338, 'Counting Bits', 'Easy', 'DP + Bits'),
        (268, 'Missing Number', 'Easy', 'XOR'),
        (371, 'Sum of Two Integers', 'Medium', 'Bit Manipulation'),
        (78, 'Subsets', 'Medium', 'Bitmasking'),
        (90, 'Subsets II', 'Medium', 'Bitmasking'),
    ],
    'trees': [
        (144, 'Binary Tree Preorder Traversal', 'Easy', 'Traversal'),
        (94, 'Binary Tree Inorder Traversal', 'Easy', 'Traversal'),
        (145, 'Binary Tree Postorder Traversal', 'Easy', 'Traversal'),
        (102, 'Binary Tree Level Order Traversal', 'Medium', 'BFS'),
        (103, 'Binary Tree Zigzag Level Order', 'Medium', 'BFS'),
        (104, 'Maximum Depth of Binary Tree', 'Easy', 'DFS'),
        (111, 'Minimum Depth of Binary Tree', 'Easy', 'BFS'),
        (110, 'Balanced Binary Tree', 'Easy', 'DFS'),
        (112, 'Path Sum', 'Easy', 'DFS'),
        (113, 'Path Sum II', 'Medium', 'DFS'),
        (236, 'Lowest Common Ancestor', 'Medium', 'DFS'),
        (199, 'Binary Tree Right Side View', 'Medium', 'BFS'),
        (226, 'Invert Binary Tree', 'Easy', 'DFS'),
        (101, 'Symmetric Tree', 'Easy', 'DFS'),
        (543, 'Diameter of Binary Tree', 'Easy', 'DFS'),
        (124, 'Binary Tree Maximum Path Sum', 'Hard', 'DFS'),
        (98, 'Validate Binary Search Tree', 'Medium', 'DFS'),
        (105, 'Construct Binary Tree from Preorder', 'Medium', 'DFS'),
        (114, 'Flatten Binary Tree to Linked List', 'Medium', 'DFS'),
        (297, 'Serialize and Deserialize Binary Tree', 'Hard', 'DFS'),
    ],
    'heaps': [
        (215, 'Kth Largest Element', 'Medium', 'Heap'),
        (347, 'Top K Frequent Elements', 'Medium', 'Heap'),
        (973, 'K Closest Points to Origin', 'Medium', 'Heap'),
        (295, 'Find Median from Data Stream', 'Hard', 'Two Heaps'),
        (378, 'Kth Smallest Element in Matrix', 'Medium', 'Heap'),
        (23, 'Merge k Sorted Lists', 'Hard', 'Heap'),
        (767, 'Reorganize String', 'Medium', 'Heap'),
        (358, 'Rearrange String k Distance Apart', 'Hard', 'Heap'),
        (621, 'Task Scheduler', 'Medium', 'Heap'),
        (703, 'Kth Largest Element in Stream', 'Easy', 'Heap'),
    ],
    'graphs': [
        (200, 'Number of Islands', 'Medium', 'DFS/BFS'),
        (695, 'Max Area of Island', 'Medium', 'DFS'),
        (130, 'Surrounded Regions', 'Medium', 'DFS'),
        (797, 'All Paths From Source to Target', 'Medium', 'DFS'),
        (547, 'Number of Provinces', 'Medium', 'DFS/Union-Find'),
        (994, 'Rotting Oranges', 'Medium', 'BFS'),
        (207, 'Course Schedule', 'Medium', 'Topological Sort'),
        (210, 'Course Schedule II', 'Medium', 'Topological Sort'),
        (133, 'Clone Graph', 'Medium', 'DFS/BFS'),
        (743, 'Network Delay Time', 'Medium', 'Dijkstra'),
        (787, 'Cheapest Flights Within K Stops', 'Medium', 'Dijkstra'),
        (684, 'Redundant Connection', 'Medium', 'Union-Find'),
        (1584, 'Min Cost to Connect All Points', 'Medium', 'MST'),
        (399, 'Evaluate Division', 'Medium', 'DFS'),
        (127, 'Word Ladder', 'Hard', 'BFS'),
        (126, 'Word Ladder II', 'Hard', 'BFS'),
        (329, 'Longest Increasing Path in Matrix', 'Hard', 'DFS + Memo'),
        (417, 'Pacific Atlantic Water Flow', 'Medium', 'DFS'),
    ],
    'tries': [
        (208, 'Implement Trie', 'Medium', 'Trie'),
        (211, 'Design Add and Search Words', 'Medium', 'Trie'),
        (212, 'Word Search II', 'Hard', 'Trie + DFS'),
        (720, 'Longest Word in Dictionary', 'Medium', 'Trie'),
        (692, 'Top K Frequent Words', 'Medium', 'Trie + Heap'),
        (642, 'Design Search Autocomplete System', 'Hard', 'Trie'),
    ],
    'recursion_backtracking': [
        (509, 'Fibonacci Number', 'Easy', 'Recursion'),
        (70, 'Climbing Stairs', 'Easy', 'DP/Recursion'),
        (50, 'Pow(x, n)', 'Medium', 'Recursion'),
        (46, 'Permutations', 'Medium', 'Backtracking'),
        (47, 'Permutations II', 'Medium', 'Backtracking'),
        (78, 'Subsets', 'Medium', 'Backtracking'),
        (90, 'Subsets II', 'Medium', 'Backtracking'),
        (51, 'N-Queens', 'Hard', 'Backtracking'),
        (52, 'N-Queens II', 'Hard', 'Backtracking'),
        (39, 'Combination Sum', 'Medium', 'Backtracking'),
        (40, 'Combination Sum II', 'Medium', 'Backtracking'),
        (22, 'Generate Parentheses', 'Medium', 'Backtracking'),
        (131, 'Palindrome Partitioning', 'Medium', 'Backtracking'),
        (79, 'Word Search', 'Medium', 'Backtracking'),
        (37, 'Sudoku Solver', 'Hard', 'Backtracking'),
        (17, 'Letter Combinations of Phone Number', 'Medium', 'Backtracking'),
    ],
    'greedy': [
        (121, 'Best Time to Buy and Sell Stock', 'Easy', 'Greedy'),
        (55, 'Jump Game', 'Medium', 'Greedy'),
        (45, 'Jump Game II', 'Medium', 'Greedy'),
        (435, 'Non-overlapping Intervals', 'Medium', 'Greedy'),
        (452, 'Minimum Arrows to Burst Balloons', 'Medium', 'Greedy'),
        (56, 'Merge Intervals', 'Medium', 'Greedy'),
        (122, 'Best Time to Buy and Sell Stock II', 'Medium', 'Greedy'),
        (134, 'Gas Station', 'Medium', 'Greedy'),
        (135, 'Candy', 'Hard', 'Greedy'),
        (406, 'Queue Reconstruction by Height', 'Medium', 'Greedy'),
        (763, 'Partition Labels', 'Medium', 'Greedy'),
        (55, 'Jump Game', 'Medium', 'Greedy'),
    ],
    'dynamic_programming': [
        (509, 'Fibonacci Number', 'Easy', '1D DP'),
        (70, 'Climbing Stairs', 'Easy', '1D DP'),
        (746, 'Min Cost Climbing Stairs', 'Easy', '1D DP'),
        (198, 'House Robber', 'Medium', '1D DP'),
        (213, 'House Robber II', 'Medium', '1D DP'),
        (416, 'Partition Equal Subset Sum', 'Medium', 'Knapsack'),
        (494, 'Target Sum', 'Medium', 'Knapsack'),
        (474, 'Ones and Zeroes', 'Medium', 'Knapsack'),
        (1143, 'Longest Common Subsequence', 'Medium', '2D DP'),
        (583, 'Delete Operation for Two Strings', 'Medium', '2D DP'),
        (72, 'Edit Distance', 'Hard', '2D DP'),
        (300, 'Longest Increasing Subsequence', 'Medium', '1D DP'),
        (673, 'Number of Longest Increasing Subsequence', 'Medium', '1D DP'),
        (354, 'Russian Doll Envelopes', 'Hard', 'LIS'),
        (139, 'Word Break', 'Medium', 'DP'),
        (132, 'Palindrome Partitioning II', 'Hard', 'DP'),
        (322, 'Coin Change', 'Medium', 'Unbounded Knapsack'),
        (518, 'Coin Change 2', 'Medium', 'Unbounded Knapsack'),
        (312, 'Burst Balloons', 'Hard', 'Interval DP'),
        (32, 'Longest Valid Parentheses', 'Hard', 'DP'),
        (85, 'Maximal Rectangle', 'Hard', 'DP'),
        (97, 'Interleaving String', 'Medium', '2D DP'),
        (115, 'Distinct Subsequences', 'Hard', '2D DP'),
        (91, 'Decode Ways', 'Medium', '1D DP'),
        (120, 'Triangle', 'Medium', '2D DP'),
        (64, 'Minimum Path Sum', 'Medium', '2D DP'),
        (62, 'Unique Paths', 'Medium', '2D DP'),
        (63, 'Unique Paths II', 'Medium', '2D DP'),
        (221, 'Maximal Square', 'Medium', '2D DP'),
        (5, 'Longest Palindromic Substring', 'Medium', '2D DP'),
    ],
}

def create_100_day_plan():
    """Create comprehensive 100-day plan."""
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    topic_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    revision_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    interview_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: Daily Schedule
    ws1 = wb.create_sheet("Daily Schedule", 0)
    headers = ['Day', 'Date', 'Topic', 'Learning Focus', 'Repository Practice', 'LeetCode Problems', 'Difficulty', 'Pattern', 'Notes']
    ws1.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # 100-day plan structure
    start_date = datetime.now()
    day_plan = generate_day_plan()
    
    for day_num, day_data in enumerate(day_plan, 1):
        current_date = start_date + timedelta(days=day_num - 1)
        date_str = current_date.strftime("%Y-%m-%d")
        
        topic = day_data['topic']
        learning = day_data['learning']
        practice = day_data['practice']
        leetcode = day_data['leetcode']
        
        # Determine fill color
        fill = None
        if 'Revision' in topic or 'Review' in topic:
            fill = revision_fill
        elif 'Mock Interview' in topic or 'Interview' in topic:
            fill = interview_fill
        elif topic:
            fill = topic_fill
        
        # Add LeetCode problems
        if leetcode:
            for i, problem in enumerate(leetcode):
                if i == 0:
                    row = [
                        day_num,
                        date_str,
                        topic,
                        learning,
                        practice,
                        f"{problem[0]}. {problem[1]}",
                        problem[2],
                        problem[3] if len(problem) > 3 else '',
                        day_data.get('notes', '')
                    ]
                else:
                    row = [
                        '', '', '', '', '',
                        f"{problem[0]}. {problem[1]}",
                        problem[2],
                        problem[3] if len(problem) > 3 else '',
                        ''
                    ]
                
                ws1.append(row)
                
                # Apply styles
                for col in range(1, len(row) + 1):
                    cell = ws1.cell(row=ws1.max_row, column=col)
                    if fill and col <= 3:
                        cell.fill = fill
                    cell.border = border
                    if col in [1, 2, 7, 8]:
                        cell.alignment = center_align
        else:
            row = [
                day_num,
                date_str,
                topic,
                learning,
                practice,
                '',
                '',
                '',
                day_data.get('notes', '')
            ]
            ws1.append(row)
            
            # Apply styles
            for col in range(1, len(row) + 1):
                cell = ws1.cell(row=ws1.max_row, column=col)
                if fill and col <= 3:
                    cell.fill = fill
                cell.border = border
                if col in [1, 2, 7, 8]:
                    cell.alignment = center_align
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 40
    ws1.column_dimensions['F'].width = 50
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 25
    ws1.column_dimensions['I'].width = 30
    
    # Sheet 2: Topic Distribution
    ws2 = wb.create_sheet("Topic Distribution")
    headers2 = ['Topic', 'Days', 'Start Day', 'End Day', 'Total Problems', 'Easy', 'Medium', 'Hard']
    ws2.append(headers2)
    
    # Style headers
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    topic_dist = calculate_topic_distribution(day_plan)
    for topic, data in topic_dist.items():
        row = [
            topic,
            data['days'],
            data['start'],
            data['end'],
            data['total_problems'],
            data['easy'],
            data['medium'],
            data['hard']
        ]
        ws2.append(row)
        
        for col in range(1, len(row) + 1):
            cell = ws2.cell(row=ws2.max_row, column=col)
            cell.border = border
            if col in [2, 3, 4, 5, 6, 7, 8]:
                cell.alignment = center_align
    
    # Adjust column widths
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 3: LeetCode Problem List
    ws3 = wb.create_sheet("LeetCode Problems")
    headers3 = ['#', 'Title', 'Difficulty', 'Pattern', 'Topic', 'Day', 'Status']
    ws3.append(headers3)
    
    # Style headers
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Collect all problems
    all_problems = []
    for day_num, day_data in enumerate(day_plan, 1):
        if day_data['leetcode']:
            for problem in day_data['leetcode']:
                all_problems.append({
                    'num': problem[0],
                    'title': problem[1],
                    'difficulty': problem[2],
                    'pattern': problem[3] if len(problem) > 3 else '',
                    'topic': day_data['topic'],
                    'day': day_num
                })
    
    # Sort by number
    all_problems.sort(key=lambda x: x['num'])
    
    for prob in all_problems:
        row = [
            prob['num'],
            prob['title'],
            prob['difficulty'],
            prob['pattern'],
            prob['topic'],
            prob['day'],
            ''  # Status column for tracking
        ]
        ws3.append(row)
        
        # Color by difficulty
        difficulty_cell = ws3.cell(row=ws3.max_row, column=3)
        if prob['difficulty'] == 'Easy':
            difficulty_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif prob['difficulty'] == 'Medium':
            difficulty_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            difficulty_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col in range(1, len(row) + 1):
            cell = ws3.cell(row=ws3.max_row, column=col)
            cell.border = border
            if col in [1, 3, 6, 7]:
                cell.alignment = center_align
    
    # Adjust column widths
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 30
    ws3.column_dimensions['E'].width = 25
    ws3.column_dimensions['F'].width = 8
    ws3.column_dimensions['G'].width = 15
    
    # Sheet 4: Progress Tracking
    ws4 = wb.create_sheet("Progress Tracking")
    headers4 = ['Week', 'Days', 'Topics Covered', 'Problems Solved', 'Target', 'Status', 'Notes']
    ws4.append(headers4)
    
    # Style headers
    for col in range(1, len(headers4) + 1):
        cell = ws4.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Weekly progress
    weeks = []
    current_week = []
    for day_num, day_data in enumerate(day_plan, 1):
        current_week.append((day_num, day_data))
        if len(current_week) == 7 or day_num == 100:
            weeks.append(current_week)
            current_week = []
    
    for week_num, week_days in enumerate(weeks, 1):
        topics = set()
        problems = 0
        for day_num, day_data in week_days:
            if day_data['topic']:
                topics.add(day_data['topic'])
            if day_data['leetcode']:
                problems += len(day_data['leetcode'])
        
        row = [
            week_num,
            f"{week_days[0][0]}-{week_days[-1][0]}",
            len(topics),
            problems,
            f"{len(week_days) * 2}-{len(week_days) * 3}",  # Target: 2-3 problems per day
            '',  # Status
            ''   # Notes
        ]
        ws4.append(row)
        
        for col in range(1, len(row) + 1):
            cell = ws4.cell(row=ws4.max_row, column=col)
            cell.border = border
            if col in [1, 2, 3, 4, 5, 6]:
                cell.alignment = center_align
    
    # Adjust column widths
    for col in range(1, 8):
        ws4.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 5: Revision Calendar
    ws5 = wb.create_sheet("Revision Calendar")
    headers5 = ['Day', 'Date', 'Revision Topics', 'Problems to Review', 'Focus Area']
    ws5.append(headers5)
    
    # Style headers
    for col in range(1, len(headers5) + 1):
        cell = ws5.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    revision_days = [day for day in day_plan if 'Revision' in day['topic'] or 'Review' in day['topic']]
    for day_num, day_data in enumerate(day_plan, 1):
        if 'Revision' in day_data['topic'] or 'Review' in day_data['topic']:
            current_date = start_date + timedelta(days=day_num - 1)
            date_str = current_date.strftime("%Y-%m-%d")
            
            row = [
                day_num,
                date_str,
                day_data.get('revision_topics', day_data['topic']),
                day_data.get('review_problems', ''),
                day_data.get('focus', '')
            ]
            ws5.append(row)
            
            for col in range(1, len(row) + 1):
                cell = ws5.cell(row=ws5.max_row, column=col)
                cell.fill = revision_fill
                cell.border = border
                if col in [1, 2]:
                    cell.alignment = center_align
    
    # Adjust column widths
    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 40
    ws5.column_dimensions['D'].width = 40
    ws5.column_dimensions['E'].width = 30
    
    return wb

def generate_day_plan():
    """Generate 100-day plan structure."""
    plan = []
    
    # Week 1: Arrays (Days 1-7)
    plan.extend([
        {'topic': 'Arrays - Core Concepts', 'learning': 'Read arrays.py, understand array properties, time/space complexity', 'practice': 'Implement basic operations, practice prefix_sum.py', 'leetcode': LEETCODE_PROBLEMS['arrays'][:3]},
        {'topic': 'Arrays - Techniques', 'learning': 'Read sliding_window.py, two_pointers.py, kadane.py', 'practice': 'Practice all techniques, solve two_sum.py, three_sum.py', 'leetcode': LEETCODE_PROBLEMS['arrays'][3:6]},
        {'topic': 'Arrays - Algorithms', 'learning': 'Read searching and sorting algorithms', 'practice': 'Implement linear_search, binary_search, basic sorts', 'leetcode': LEETCODE_PROBLEMS['arrays'][6:9]},
        {'topic': 'Arrays - Interview Problems', 'learning': 'Review all array interview problems', 'practice': 'Solve all problems in interview_problems/', 'leetcode': LEETCODE_PROBLEMS['arrays'][9:12]},
        {'topic': 'Arrays - Advanced Problems', 'learning': 'Review advanced array patterns', 'practice': 'Solve complex array problems', 'leetcode': LEETCODE_PROBLEMS['arrays'][12:15]},
        {'topic': 'Arrays - Mixed Practice', 'learning': 'Review array concepts', 'practice': 'Solve missed problems, practice techniques', 'leetcode': LEETCODE_PROBLEMS['arrays'][15:18]},
        {'topic': 'Week 1 Review - Arrays', 'learning': 'Comprehensive review of arrays', 'practice': 'Solve all array problems from repository', 'leetcode': LEETCODE_PROBLEMS['arrays'][18:20] + LEETCODE_PROBLEMS['binary_search'][:2], 'revision_topics': 'Arrays: all techniques and algorithms', 'focus': 'Pattern recognition'},
    ])
    
    # Week 2: Linked Lists & Strings (Days 8-14)
    plan.extend([
        {'topic': 'Linked Lists - Core', 'learning': 'Read singly_linked_list.py, doubly_linked_list.py', 'practice': 'Implement all basic operations, reversal, cycle detection', 'leetcode': LEETCODE_PROBLEMS['linked_lists'][:3]},
        {'topic': 'Linked Lists - Techniques', 'learning': 'Read slow_fast_pointer.py, dummy_node_technique.py', 'practice': 'Practice all techniques, solve interview problems', 'leetcode': LEETCODE_PROBLEMS['linked_lists'][3:6]},
        {'topic': 'Linked Lists - Problems', 'learning': 'Review linked list patterns', 'practice': 'Solve all linked list interview problems', 'leetcode': LEETCODE_PROBLEMS['linked_lists'][6:9]},
        {'topic': 'Strings - Core & Techniques', 'learning': 'Read strings.py, all techniques', 'practice': 'Practice string operations, sliding window, two pointers', 'leetcode': LEETCODE_PROBLEMS['strings'][:3]},
        {'topic': 'Strings - Algorithms', 'learning': 'Read KMP, Rabin-Karp algorithms', 'practice': 'Implement string algorithms, solve interview problems', 'leetcode': LEETCODE_PROBLEMS['strings'][3:6]},
        {'topic': 'Strings - Problems', 'learning': 'Review string patterns', 'practice': 'Solve all string interview problems', 'leetcode': LEETCODE_PROBLEMS['strings'][6:9]},
        {'topic': 'Week 2 Review - Linked Lists & Strings', 'learning': 'Review linked lists and strings', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['linked_lists'][9:12] + LEETCODE_PROBLEMS['strings'][9:12], 'revision_topics': 'Linked Lists, Strings', 'focus': 'Implementation skills'},
    ])
    
    # Week 3: Binary Search & Sorting (Days 15-21)
    plan.extend([
        {'topic': 'Binary Search - Fundamentals', 'learning': 'Read iterative.py, recursive.py', 'practice': 'Implement all binary search variants', 'leetcode': LEETCODE_PROBLEMS['binary_search'][:3]},
        {'topic': 'Binary Search - Advanced', 'learning': 'Read rotated_array.py, binary_search_on_answer.py', 'practice': 'Practice rotated arrays, peak finding', 'leetcode': LEETCODE_PROBLEMS['binary_search'][3:6]},
        {'topic': 'Binary Search - Problems', 'learning': 'Review binary search patterns', 'practice': 'Solve all binary search problems', 'leetcode': LEETCODE_PROBLEMS['binary_search'][6:9]},
        {'topic': 'Sorting - Basic Algorithms', 'learning': 'Read bubble_sort.py, selection_sort.py, insertion_sort.py', 'practice': 'Implement all basic sorts', 'leetcode': LEETCODE_PROBLEMS['sorting'][:3]},
        {'topic': 'Sorting - Advanced', 'learning': 'Read merge_sort.py, quick_sort.py', 'practice': 'Implement advanced sorts, practice applications', 'leetcode': LEETCODE_PROBLEMS['sorting'][3:6]},
        {'topic': 'Sorting - Problems', 'learning': 'Review sorting patterns', 'practice': 'Solve all sorting interview problems', 'leetcode': LEETCODE_PROBLEMS['sorting'][6:9]},
        {'topic': 'Week 3 Review - Binary Search & Sorting', 'learning': 'Review binary search and sorting', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['binary_search'][9:12] + LEETCODE_PROBLEMS['sorting'][9:12], 'revision_topics': 'Binary Search, Sorting', 'focus': 'Algorithm efficiency'},
    ])
    
    # Week 4: Stacks, Queues, Hash Tables (Days 22-28)
    plan.extend([
        {'topic': 'Stacks - Core & Applications', 'learning': 'Read stacks.py, all applications', 'practice': 'Implement stack, practice applications', 'leetcode': LEETCODE_PROBLEMS['stacks'][:3]},
        {'topic': 'Stacks - Interview Problems', 'learning': 'Review stack patterns', 'practice': 'Solve all stack interview problems', 'leetcode': LEETCODE_PROBLEMS['stacks'][3:6]},
        {'topic': 'Queues - Core & Applications', 'learning': 'Read queues.py, all types and applications', 'practice': 'Implement all queue types, practice applications', 'leetcode': LEETCODE_PROBLEMS['queues'][:3]},
        {'topic': 'Queues - Problems', 'learning': 'Review queue patterns', 'practice': 'Solve all queue interview problems', 'leetcode': LEETCODE_PROBLEMS['queues'][3:6]},
        {'topic': 'Hash Tables - Core Concepts', 'learning': 'Read hash_tables.py, hashing.py, collision_handling.py', 'practice': 'Understand hash functions, collision resolution', 'leetcode': LEETCODE_PROBLEMS['hash_tables'][:3]},
        {'topic': 'Hash Tables - Applications', 'learning': 'Read all hash table applications', 'practice': 'Practice frequency maps, prefix sum optimization', 'leetcode': LEETCODE_PROBLEMS['hash_tables'][3:6]},
        {'topic': 'Week 4 Review - Stacks, Queues, Hash Tables', 'learning': 'Review all linear data structures', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['stacks'][6:9] + LEETCODE_PROBLEMS['hash_tables'][6:9], 'revision_topics': 'Stacks, Queues, Hash Tables', 'focus': 'Data structure selection'},
    ])
    
    # Week 5: Bit Manipulation & Trees (Days 29-35)
    plan.extend([
        {'topic': 'Bit Manipulation - Core', 'learning': 'Read bits.py, all techniques', 'practice': 'Practice all bit operations manually', 'leetcode': LEETCODE_PROBLEMS['bit_manipulation'][:3]},
        {'topic': 'Bit Manipulation - Algorithms', 'learning': 'Read all bit manipulation algorithms', 'practice': 'Implement algorithms, solve interview problems', 'leetcode': LEETCODE_PROBLEMS['bit_manipulation'][3:6]},
        {'topic': 'Bit Manipulation - Problems', 'learning': 'Review bit manipulation patterns', 'practice': 'Solve all bit manipulation problems', 'leetcode': LEETCODE_PROBLEMS['bit_manipulation'][6:9]},
        {'topic': 'Trees - Core & Traversals', 'learning': 'Read trees.py, all traversals', 'practice': 'Implement all traversals (iterative & recursive)', 'leetcode': LEETCODE_PROBLEMS['trees'][:3]},
        {'topic': 'Trees - Search Algorithms', 'learning': 'Read DFS.py, BFS.py', 'practice': 'Practice DFS and BFS, level-order traversal', 'leetcode': LEETCODE_PROBLEMS['trees'][3:6]},
        {'topic': 'Trees - Properties', 'learning': 'Read height.py, diameter.py, lca.py', 'practice': 'Implement all tree properties', 'leetcode': LEETCODE_PROBLEMS['trees'][6:9]},
        {'topic': 'Week 5 Review - Bit Manipulation & Trees', 'learning': 'Review bit manipulation and trees', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['bit_manipulation'][9:12] + LEETCODE_PROBLEMS['trees'][9:12], 'revision_topics': 'Bit Manipulation, Trees', 'focus': 'Tree algorithms'},
    ])
    
    # Week 6: Trees Continued & Heaps (Days 36-42)
    plan.extend([
        {'topic': 'Trees - Interview Problems', 'learning': 'Review all tree interview problems', 'practice': 'Solve all tree interview problems', 'leetcode': LEETCODE_PROBLEMS['trees'][12:15]},
        {'topic': 'Trees - Advanced Problems', 'learning': 'Review advanced tree patterns', 'practice': 'Solve complex tree problems', 'leetcode': LEETCODE_PROBLEMS['trees'][15:18]},
        {'topic': 'Trees - BST Problems', 'learning': 'Review BST-specific problems', 'practice': 'Solve BST validation, construction problems', 'leetcode': LEETCODE_PROBLEMS['trees'][18:20]},
        {'topic': 'Heaps - Core & Types', 'learning': 'Read heaps.py, min_heap.py, max_heap.py', 'practice': 'Implement min and max heaps', 'leetcode': LEETCODE_PROBLEMS['heaps'][:3]},
        {'topic': 'Heaps - Algorithms', 'learning': 'Read heap_sort.py, all algorithms', 'practice': 'Implement heap sort, practice applications', 'leetcode': LEETCODE_PROBLEMS['heaps'][3:6]},
        {'topic': 'Heaps - Problems', 'learning': 'Review heap patterns', 'practice': 'Solve all heap interview problems', 'leetcode': LEETCODE_PROBLEMS['heaps'][6:9]},
        {'topic': 'Week 6 Review - Trees & Heaps', 'learning': 'Review trees and heaps', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['trees'][:2] + LEETCODE_PROBLEMS['heaps'][9:10], 'revision_topics': 'Trees, Heaps', 'focus': 'Priority queue applications'},
    ])
    
    # Week 7: Graphs (Days 43-49)
    plan.extend([
        {'topic': 'Graphs - Core & Representations', 'learning': 'Read graphs.py, directed_undirected.py, weighted_unweighted.py', 'practice': 'Implement graph representations', 'leetcode': LEETCODE_PROBLEMS['graphs'][:3]},
        {'topic': 'Graphs - DFS & BFS', 'learning': 'Read dfs.py, bfs.py', 'practice': 'Implement DFS and BFS', 'leetcode': LEETCODE_PROBLEMS['graphs'][3:6]},
        {'topic': 'Graphs - Algorithms (Part 1)', 'learning': 'Read dijkstra.py, union_find.py', 'practice': 'Implement Dijkstra\'s, Union-Find', 'leetcode': LEETCODE_PROBLEMS['graphs'][6:9]},
        {'topic': 'Graphs - Algorithms (Part 2)', 'learning': 'Read prim_mst.py, kruskal_mst.py', 'practice': 'Implement Prim\'s and Kruskal\'s', 'leetcode': LEETCODE_PROBLEMS['graphs'][9:12]},
        {'topic': 'Graphs - Interview Problems', 'learning': 'Review graph interview problems', 'practice': 'Solve all graph interview problems', 'leetcode': LEETCODE_PROBLEMS['graphs'][12:15]},
        {'topic': 'Graphs - Advanced Problems', 'learning': 'Review advanced graph patterns', 'practice': 'Solve complex graph problems', 'leetcode': LEETCODE_PROBLEMS['graphs'][15:18]},
        {'topic': 'Week 7 Review - Graphs', 'learning': 'Comprehensive graph review', 'practice': 'Solve missed graph problems', 'leetcode': [], 'revision_topics': 'Graphs: all algorithms', 'focus': 'Graph algorithm selection'},
    ])
    
    # Week 8: Tries & Recursion (Days 50-56)
    plan.extend([
        {'topic': 'Tries - Core & Operations', 'learning': 'Read trie.py, insert.py, search.py', 'practice': 'Implement trie from scratch', 'leetcode': LEETCODE_PROBLEMS['tries'][:3]},
        {'topic': 'Tries - Applications', 'learning': 'Read autocomplete.py', 'practice': 'Implement autocomplete system', 'leetcode': LEETCODE_PROBLEMS['tries'][3:6]},
        {'topic': 'Recursion - Core Concepts', 'learning': 'Read recursion.py', 'practice': 'Implement all recursion examples', 'leetcode': LEETCODE_PROBLEMS['recursion_backtracking'][:3]},
        {'topic': 'Backtracking - Core Techniques', 'learning': 'Read all backtracking techniques', 'practice': 'Implement N-Queens, subsets, permutations', 'leetcode': LEETCODE_PROBLEMS['recursion_backtracking'][3:6]},
        {'topic': 'Backtracking - Applications', 'learning': 'Read sudoku_solver.py, rat_in_maze.py', 'practice': 'Implement all backtracking applications', 'leetcode': LEETCODE_PROBLEMS['recursion_backtracking'][6:9]},
        {'topic': 'Backtracking - Problems', 'learning': 'Review backtracking patterns', 'practice': 'Solve all backtracking interview problems', 'leetcode': LEETCODE_PROBLEMS['recursion_backtracking'][9:12]},
        {'topic': 'Week 8 Review - Tries & Recursion', 'learning': 'Review tries and recursion/backtracking', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['recursion_backtracking'][12:16], 'revision_topics': 'Tries, Recursion, Backtracking', 'focus': 'Recursive thinking'},
    ])
    
    # Week 9: Greedy Algorithms (Days 57-63)
    plan.extend([
        {'topic': 'Greedy - Core Concepts', 'learning': 'Read greedy.py', 'practice': 'Understand greedy choice property', 'leetcode': LEETCODE_PROBLEMS['greedy'][:3]},
        {'topic': 'Greedy - Applications', 'learning': 'Read interval_scheduling.py, huffman_encoding.py', 'practice': 'Implement all greedy applications', 'leetcode': LEETCODE_PROBLEMS['greedy'][3:6]},
        {'topic': 'Greedy - Problems', 'learning': 'Review greedy patterns', 'practice': 'Solve all greedy interview problems', 'leetcode': LEETCODE_PROBLEMS['greedy'][6:9]},
        {'topic': 'Greedy - Advanced Problems', 'learning': 'Review advanced greedy problems', 'practice': 'Solve complex greedy problems', 'leetcode': LEETCODE_PROBLEMS['greedy'][9:12]},
        {'topic': 'Revision - Arrays & Strings', 'learning': 'Review arrays and strings', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['arrays'][:3] + LEETCODE_PROBLEMS['strings'][:3], 'revision_topics': 'Arrays, Strings', 'focus': 'Pattern recognition'},
        {'topic': 'Revision - Linked Lists & Binary Search', 'learning': 'Review linked lists and binary search', 'practice': 'Solve missed problems', 'leetcode': LEETCODE_PROBLEMS['linked_lists'][:3] + LEETCODE_PROBLEMS['binary_search'][:3], 'revision_topics': 'Linked Lists, Binary Search', 'focus': 'Implementation'},
        {'topic': 'Week 9 Review - Greedy & Revisions', 'learning': 'Review greedy and previous topics', 'practice': 'Mixed practice', 'leetcode': [], 'revision_topics': 'All topics covered so far', 'focus': 'Mixed problem solving'},
    ])
    
    # Week 10: Dynamic Programming - Part 1 (Days 64-70)
    plan.extend([
        {'topic': 'DP - Core Concepts & Approaches', 'learning': 'Read dp.py, all approaches', 'practice': 'Compare recursive vs memoized vs tabulated', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][:3]},
        {'topic': 'DP - Classic Problems (1D)', 'learning': 'Read fibonacci.py, understand 1D DP', 'practice': 'Implement Fibonacci, climbing stairs variations', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][3:6]},
        {'topic': 'DP - Classic Problems (Knapsack)', 'learning': 'Read knapsack.py', 'practice': 'Implement 0/1 knapsack and variations', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][6:9]},
        {'topic': 'DP - Classic Problems (2D)', 'learning': 'Read lcs.py, understand 2D DP', 'practice': 'Implement LCS and variations', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][9:12]},
        {'topic': 'DP - Patterns (Subsequences)', 'learning': 'Read subsequences.py', 'practice': 'Implement LIS, distinct subsequences', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][12:15]},
        {'topic': 'DP - Patterns (Partitioning)', 'learning': 'Read partition.py', 'practice': 'Implement partitioning problems', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][15:18]},
        {'topic': 'Week 10 Review - DP Basics', 'learning': 'Review DP fundamentals', 'practice': 'Solve missed DP problems', 'leetcode': [], 'revision_topics': 'DP: 1D, 2D, Knapsack', 'focus': 'DP pattern recognition'},
    ])
    
    # Week 11: Dynamic Programming - Part 2 (Days 71-77)
    plan.extend([
        {'topic': 'DP - Interview Problems (Part 1)', 'learning': 'Read min_edit_distance.py, coin_change.py', 'practice': 'Implement edit distance, coin change', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][18:21]},
        {'topic': 'DP - Interview Problems (Part 2)', 'learning': 'Read matrix_chain_multiplication.py', 'practice': 'Implement matrix chain multiplication', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][21:24]},
        {'topic': 'DP - Advanced Patterns', 'learning': 'Review advanced DP patterns', 'practice': 'Practice complex DP problems', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][24:27]},
        {'topic': 'DP - Grid Problems', 'learning': 'Review DP on grids', 'practice': 'Solve grid DP problems', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][27:30]},
        {'topic': 'DP - String Problems', 'learning': 'Review DP on strings', 'practice': 'Solve string DP problems', 'leetcode': [LEETCODE_PROBLEMS['dynamic_programming'][9], LEETCODE_PROBLEMS['dynamic_programming'][10], LEETCODE_PROBLEMS['dynamic_programming'][11]]},
        {'topic': 'DP - Mixed Practice', 'learning': 'Review all DP patterns', 'practice': 'Solve mixed DP problems', 'leetcode': []},
        {'topic': 'Week 11 Review - DP Advanced', 'learning': 'Comprehensive DP review', 'practice': 'Solve missed DP problems', 'leetcode': [], 'revision_topics': 'DP: All patterns', 'focus': 'DP optimization'},
    ])
    
    # Week 12: Comprehensive Review & Mock Interviews (Days 78-84)
    plan.extend([
        {'topic': 'Revision - Data Structures', 'learning': 'Review all data structures', 'practice': 'Solve repository problems', 'leetcode': LEETCODE_PROBLEMS['arrays'][:2] + LEETCODE_PROBLEMS['linked_lists'][:2] + LEETCODE_PROBLEMS['trees'][:2], 'revision_topics': 'Arrays, Linked Lists, Trees, Stacks, Queues', 'focus': 'Data structure selection'},
        {'topic': 'Revision - Algorithms', 'learning': 'Review all algorithms', 'practice': 'Solve repository problems', 'leetcode': LEETCODE_PROBLEMS['sorting'][:2] + LEETCODE_PROBLEMS['binary_search'][:2] + LEETCODE_PROBLEMS['graphs'][:2], 'revision_topics': 'Sorting, Searching, Graph Algorithms', 'focus': 'Algorithm efficiency'},
        {'topic': 'Revision - Problem Solving Techniques', 'learning': 'Review DP, Greedy, Backtracking', 'practice': 'Solve repository problems', 'leetcode': LEETCODE_PROBLEMS['dynamic_programming'][:2] + LEETCODE_PROBLEMS['greedy'][:2] + LEETCODE_PROBLEMS['recursion_backtracking'][:2], 'revision_topics': 'DP, Greedy, Backtracking', 'focus': 'Technique selection'},
        {'topic': 'Mock Interview Practice - Day 1', 'learning': 'Timed problem solving', 'practice': 'Solve 3-4 problems (45 min each)', 'leetcode': [], 'notes': 'Focus: Arrays, Strings, Linked Lists'},
        {'topic': 'Mock Interview Practice - Day 2', 'learning': 'Timed problem solving', 'practice': 'Solve 3-4 problems (45 min each)', 'leetcode': [], 'notes': 'Focus: Trees, Graphs'},
        {'topic': 'Mock Interview Practice - Day 3', 'learning': 'Timed problem solving', 'practice': 'Solve 3-4 problems (45 min each)', 'leetcode': [], 'notes': 'Focus: DP, Greedy'},
        {'topic': 'Week 12 Review - Comprehensive', 'learning': 'Review all topics', 'practice': 'Solve weak areas', 'leetcode': [], 'revision_topics': 'All topics', 'focus': 'Weak area improvement'},
    ])
    
    # Week 13: Final Preparation (Days 85-91)
    plan.extend([
        {'topic': 'Mixed Practice - Easy & Medium', 'learning': 'Solve problems from all topics', 'practice': 'Focus on medium problems', 'leetcode': []},
        {'topic': 'Mixed Practice - Medium Focus', 'learning': 'Solve medium problems', 'practice': 'Time yourself', 'leetcode': []},
        {'topic': 'Pattern Recognition Practice', 'learning': 'Identify patterns quickly', 'practice': 'Solve problems by pattern', 'leetcode': []},
        {'topic': 'Mock Interview - Full Session', 'learning': 'Full interview simulation', 'practice': '2 problems in 90 minutes', 'leetcode': [], 'notes': 'Include explanation practice'},
        {'topic': 'Weak Area Deep Dive', 'learning': 'Focus on weak topics', 'practice': 'Solve problems from weak areas', 'leetcode': []},
        {'topic': 'Revision - All Topics', 'learning': 'Quick review of all topics', 'practice': 'Solve repository problems', 'leetcode': []},
        {'topic': 'Week 13 Review', 'learning': 'Assess progress', 'practice': 'Solve missed problems', 'leetcode': [], 'revision_topics': 'All topics', 'focus': 'Confidence building'},
    ])
    
    # Week 14: Final Sprint (Days 92-100)
    plan.extend([
        {'topic': 'Final Review - Arrays to Hash Tables', 'learning': 'Review first half of syllabus', 'practice': 'Solve problems', 'leetcode': []},
        {'topic': 'Final Review - Trees to DP', 'learning': 'Review second half of syllabus', 'practice': 'Solve problems', 'leetcode': []},
        {'topic': 'Mock Interview - Day 1', 'learning': 'Full interview simulation', 'practice': '2-3 problems with explanation', 'leetcode': []},
        {'topic': 'Mock Interview - Day 2', 'learning': 'Full interview simulation', 'practice': '2-3 problems with explanation', 'leetcode': []},
        {'topic': 'Pattern Practice', 'learning': 'Practice recognizing patterns', 'practice': 'Solve problems by pattern', 'leetcode': []},
        {'topic': 'Time Management Practice', 'learning': 'Practice solving under time', 'practice': 'Time yourself on problems', 'leetcode': []},
        {'topic': 'Final Assessment - Day 1', 'learning': 'Solve 5-6 problems', 'practice': 'Mix of all topics', 'leetcode': []},
        {'topic': 'Final Assessment - Day 2', 'learning': 'Solve 5-6 problems', 'practice': 'Mix of all topics', 'leetcode': []},
        {'topic': 'Final Review & Reflection', 'learning': 'Review what you learned', 'practice': 'Identify strengths/weaknesses', 'leetcode': []},
    ])
    
    # Fill remaining days with mixed practice
    while len(plan) < 100:
        plan.append({
            'topic': 'Mixed Practice',
            'learning': 'Solve problems from all topics',
            'practice': 'Focus on weak areas',
            'leetcode': []
        })
    
    return plan[:100]

def calculate_topic_distribution(day_plan):
    """Calculate topic distribution statistics."""
    topic_stats = {}
    
    for day_num, day_data in enumerate(day_plan, 1):
        topic = day_data['topic']
        if not topic or 'Practice' in topic or 'Review' in topic or 'Mock' in topic:
            continue
        
        # Extract main topic
        main_topic = topic.split(' -')[0].split(' (')[0].strip()
        
        if main_topic not in topic_stats:
            topic_stats[main_topic] = {
                'days': 0,
                'start': day_num,
                'end': day_num,
                'total_problems': 0,
                'easy': 0,
                'medium': 0,
                'hard': 0
            }
        
        topic_stats[main_topic]['days'] += 1
        topic_stats[main_topic]['end'] = day_num
        
        if day_data['leetcode']:
            for problem in day_data['leetcode']:
                topic_stats[main_topic]['total_problems'] += 1
                if problem[2] == 'Easy':
                    topic_stats[main_topic]['easy'] += 1
                elif problem[2] == 'Medium':
                    topic_stats[main_topic]['medium'] += 1
                else:
                    topic_stats[main_topic]['hard'] += 1
    
    return topic_stats

if __name__ == "__main__":
    print("Generating 100-Day DSA Preparation Plan...")
    wb = create_100_day_plan()
    
    filename = "100_DAY_PREPARATION_PLAN.xlsx"
    wb.save(filename)
    print(f"\n✅ Excel file created: {filename}")
    print("\nSheets included:")
    print("  1. Daily Schedule - Complete 100-day plan")
    print("  2. Topic Distribution - Statistics by topic")
    print("  3. LeetCode Problems - All problems sorted by number")
    print("  4. Progress Tracking - Weekly progress tracker")
    print("  5. Revision Calendar - All revision days")
    
    # Print summary statistics
    day_plan = generate_day_plan()
    total_problems = sum(len(day['leetcode']) for day in day_plan if day['leetcode'])
    topic_dist = calculate_topic_distribution(day_plan)
    
    print(f"\n📊 Summary Statistics:")
    print(f"  Total Days: 100")
    print(f"  Total LeetCode Problems: {total_problems}")
    print(f"  Topics Covered: {len(topic_dist)}")
    print(f"\nTopic Distribution:")
    for topic, stats in sorted(topic_dist.items()):
        print(f"  {topic}: {stats['days']} days, {stats['total_problems']} problems")

